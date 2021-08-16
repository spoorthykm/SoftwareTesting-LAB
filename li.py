import re
import openpyxl
from openpyxl import Workbook,load_workbook

wb = load_workbook("license.xlsx")
ws = wb.active

rows = ws.max_row
columns = ws.max_column
for i in range(1,rows+1):
    cell_obj = ws.cell(row=i,column=2)
    key = str(cell_obj.value)
    res = re.search(r'ABC[A-Za-z0-9@#$%^&*()]{10}XYZ',key)
    if res:
        print(ws.cell(row=i,column=1).value,"           ",res.group())