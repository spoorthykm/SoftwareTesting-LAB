import os
from openpyxl import Workbook,load_workbook

if os.path.exists("spoo.xlsx"):
    wb=load_workbook("spoo.xlsx")
    ws=wb.worksheets[0]
else:
    wb= Workbook()
    ws=wb.active
ws.append(["Name","USN","Marks1","Marks2","Marks3"])
for i in range (2):
    print("Enter Name USN Marks of 3 subject")
    data=input().split(" ")
    ws.append([data[0],data[1],data[2],data[3],data[4]])
wb.save("spoo.xlsx")
os.system("spoo.xlsx")
