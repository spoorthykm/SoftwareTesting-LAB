import re
import openpyxl
from openpyxl import load_workbook

wb=load_workbook("license.xlsx")
ws=wb.active

row=ws.max_row
col=ws.max_column
for i in range(1,row+1):
    cell_obj= ws.cell(i,2)
    key=str(cell_obj.value)
    res=re.search(r'ABC[a-zA-Z0-9@#$*()]{10}XYZ',key)
    if res:
        print(res.group())
